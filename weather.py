from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
import logging
from typing import TypeAlias

import requests
from openpyxl import load_workbook

from constants import AMBER_VAL, COUNTY, GOOD_DAY_THRESHOLD, GREEN_VAL, MET_IE_FORECAST_CUR

DIVIDER = "+" + "-" * 11 + "+" + "-" * 10 + "+" + "-" * 13 + "+" + "-" * 8 + "+"
TableRow: TypeAlias = tuple[str, str, int, str]
PeriodForecast: TypeAlias = dict[str, tuple[int, str]]
InverterPlan: TypeAlias = dict[str, str]


class SolarForecast:
    """Download and interpret solar forecast data for one county."""

    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self.table_data: list[TableRow] = []
        self.__get_solar_forecast_table(COUNTY)

    @staticmethod
    def __get_today() -> str:
        today_day = datetime.now().strftime("%A").upper()
        return today_day[:3].capitalize()

    @staticmethod
    def __status_from_value(value: int) -> str:
        """Convert a numeric solar value into Red, Amber, or Green status."""
        if value < 200:
            return "Red"
        if value <= 400:
            return "Amber"
        return "Green"

    @staticmethod
    def __parse_csv_rows(content: bytes) -> list[dict[str, str]]:
        """Parse bytes as CSV and return non-empty row dictionaries."""
        csv_text = content.decode("utf-8")
        reader = csv.DictReader(StringIO(csv_text, newline=""))
        return [dict(row) for row in reader if any((value or "").strip() for value in row.values())]

    @staticmethod
    def __parse_excel_rows(content: bytes) -> list[dict[str, str]]:
        """Parse bytes as Excel and map rows using first row as headers."""
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active

        row_iterator = sheet.iter_rows(values_only=True)
        headers_row = next(row_iterator, None)
        if headers_row is None:
            return []

        headers: list[str] = [str(header).strip() for header in headers_row]
        data_rows: list[dict[str, str]] = []
        for row in row_iterator:
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue
            data_rows.append(dict(zip(headers, values)))
        return data_rows

    def __get_rows(self, content: bytes) -> list[dict[str, str]]:
        """Read response bytes as CSV first, then fall back to Excel."""
        try:
            rows = self.__parse_csv_rows(content)
            if rows and any("SOLAR" in key for key in rows[0].keys()):
                return rows
        except (UnicodeDecodeError, csv.Error):
            pass

        return self.__parse_excel_rows(content)

    def __get_solar_forecast_table(self, county: str) -> None:
        # Download the latest forecast CSV from ESB Networks.
        response = requests.get(MET_IE_FORECAST_CUR, timeout=30)
        response.raise_for_status()

        # Parse the response rows, supporting both CSV and Excel formats.
        rows: list[dict[str, str]] = self.__get_rows(response.content)
        if not rows:
            raise ValueError("Forecast data did not contain any rows")

        # Find the county row by checking the first column value case-insensitively.
        first_column_name = next(iter(rows[0].keys()), None)
        if first_column_name is None:
            raise ValueError("Forecast data did not include headers")

        county_row: dict[str, str] | None = next(
            (
                row for row in rows
                if row.get(first_column_name, "").strip().upper() == county.strip().upper()
            ),
            None,
        )
        if county_row is None:
            raise ValueError(f"County '{county}' was not found in the forecast data")

        # Keep only solar columns, such as SOLAR-MON-MORNING.
        solar_columns: list[str] = [col for col in rows[0].keys() if "SOLAR" in col]
        periods: list[str] = [col.replace("SOLAR-", "") for col in solar_columns]

        # Convert each forecast column into a printable table row.
        for col, period in zip(solar_columns, periods):
            raw_value = county_row.get(col)
            if raw_value is None:
                continue
            value = int(raw_value)
            status = self.__status_from_value(value)
            day, time = period.split('-')
            self.table_data.append((day, time, value, status))

        # Print a simple ASCII table and separate each day for readability.
        header = f"| {'Day':^9} | {'Period':^8} | {'Solar Value':^11} | {'Status':^6} |"
        self.logger.info(DIVIDER)
        self.logger.info(header)
        self.logger.info(DIVIDER)

        previous_day: str | None = None
        for row in self.table_data:
            day, time, value, status = row
            if previous_day and day != previous_day:
                self.logger.info(DIVIDER)
            self.logger.info(f"| {day:^9} | {time:^8} | {value:^11} | {status:^6} |")
            previous_day = day

        self.logger.info(DIVIDER)

    def get_todays_solar_values(self) -> list[str]:
        today_day = SolarForecast.__get_today()

        # Print today's rows and convert values to short status codes.
        self.logger.info(f"Solar Values for today ({today_day}):")

        values: list[str] = []
        self.logger.info(DIVIDER)
        for row in self.table_data:
            day, time, value, status = row
            if day == today_day and time != "NIGHT":
                self.logger.info(f"| {day:^9} | {time:^8} | {value:^11} | {status:^6} |")
                if status == "Red":
                    values.append("R")
                elif status == "Amber":
                    values.append("A")
                else:
                    values.append("G")
        self.logger.info(DIVIDER)

        return values

    def get_todays_period_forecast(self) -> PeriodForecast:
        """Return today's Morn/Aftn/Eve forecast values and status by period."""
        today_day = SolarForecast.__get_today()
        period_forecast: PeriodForecast = {}

        for day, time, value, status in self.table_data:
            if day != today_day or time == "NIGHT":
                continue
            period_forecast[time] = (value, status)

        return period_forecast

    def get_simple_inverter_plan(self) -> InverterPlan:
        """Create a simple planning hint for charge/discharge by daytime period."""
        forecast = self.get_todays_period_forecast()
        plan: InverterPlan = {}

        # This is a simple rule-based plan that can be tuned with your tariff and battery size.
        for period in ("Morn", "Aftn", "Eve"):
            value_status = forecast.get(period)
            if value_status is None:
                plan[period] = "No forecast available"
                continue

            _, status = value_status
            if status == "Green":
                plan[period] = "Prefer discharge first to create battery headroom and reduce clipping risk"
            elif status == "Amber":
                plan[period] = "Use balanced mode with light charging/discharging"
            else:
                plan[period] = "Avoid discharge where possible; keep battery energy for home use"

        return plan

    def is_good_day(self) -> bool:
        # Use weighted scoring to decide if today is a good solar day.
        today: list[str] = self.get_todays_solar_values()
        good_periods = sum(GREEN_VAL for v in today if v == "G")
        good_periods += sum(AMBER_VAL for v in today if v == "A")
        return good_periods >= GOOD_DAY_THRESHOLD


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    log = logging.getLogger(__name__)
    solar_forecast = SolarForecast(log)
    todays_values = solar_forecast.get_todays_solar_values()
    todays_periods = solar_forecast.get_todays_period_forecast()
    inverter_plan = solar_forecast.get_simple_inverter_plan()
    log.info(f"Today's forecast: {todays_values}")
    log.info(f"Today's period forecast: {todays_periods}")
    log.info(f"Simple inverter plan: {inverter_plan}")
    log.info(f"A good day? {solar_forecast.is_good_day()}")
